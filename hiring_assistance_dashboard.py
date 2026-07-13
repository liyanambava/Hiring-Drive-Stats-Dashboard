"""
Hiring Assistance Dashboard
============================
A Streamlit app for CV-screening trackers (Shell / Qlik Sense style sheets).

Run with:  streamlit run hiring_assistance_dashboard.py
"""

import hashlib
import re
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Page config & design tokens
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Hiring Assistance Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

NAVY      = "#0F2D5E"
BLUE      = "#1F6FBB"
LIGHT_BLUE= "#EAF2FB"
TEAL      = "#0D8A72"
RED       = "#C0392B"
AMBER     = "#D4870A"
BG        = "#F4F7FB"
SURFACE   = "#FFFFFF"
BORDER    = "#DDE3EE"
TEXT      = "#1A1A2E"
MUTED     = "#6B7A99"

# Chart palette
PALETTE   = [BLUE, TEAL, AMBER, "#7B61FF", "#E87040", "#2EBFA5", "#C0392B"]

st.markdown(
    f"""
    <style>
        /* ── Global background ── */
        .stApp {{ background-color: {BG}; }}

        /* ── Hide default Streamlit chrome ── */
        #footer {{ visibility: hidden; }}

        /* ── Sidebar ── */
        section[data-testid="stSidebar"] {{
            background-color: {NAVY};
        }}
        section[data-testid="stSidebar"] * {{
            color: #E8EEF8 !important;
        }}
        section[data-testid="stSidebar"] .stSlider > div > div > div {{
            background: {BLUE} !important;
        }}
        section[data-testid="stSidebar"] label {{
            font-weight: 600 !important;
            font-size: 0.82rem !important;
            letter-spacing: 0.04em !important;
            text-transform: uppercase !important;
        }}

        /* ── File uploader: hint, button & uploaded file names/sizes ── */
        section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzoneInstructions"] span,
        section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzoneInstructions"] small,
        section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzoneInstructions"] p {{
            color: #000000 !important;
        }}
        section[data-testid="stSidebar"] [data-testid="stFileUploader"] button span {{
            color: #000000 !important;
        }}
        /* Uploaded file name and size in the file list */
        section[data-testid="stSidebar"] [data-testid="stUploadedFile"] span,
        section[data-testid="stSidebar"] [data-testid="stUploadedFile"] p,
        section[data-testid="stSidebar"] [data-testid="stUploadedFile"] small,
        section[data-testid="stSidebar"] [data-testid="stUploadedFileData"] span,
        section[data-testid="stSidebar"] [data-testid="stUploadedFileData"] p {{
            color: #000000 !important;
        }}
        /* The delete (×) button on each uploaded file */
        section[data-testid="stSidebar"] [data-testid="stFileUploaderDeleteBtn"] button span {{
            color: #000000 !important;
        }}

        /* ── Banner ── */
        .dash-banner {{
            background: linear-gradient(135deg, {NAVY} 0%, {BLUE} 100%);
            border-radius: 14px;
            padding: 28px 36px 22px 36px;
            margin-bottom: 24px;
        }}
        .dash-banner h1 {{
            color: #FFFFFF;
            font-size: 1.95rem;
            font-weight: 800;
            margin: 0 0 4px 0;
            letter-spacing: -0.02em;
        }}
        .dash-banner p {{
            color: #B8CFEE;
            font-size: 0.92rem;
            margin: 0;
        }}

        /* ── KPI cards ── */
        .kpi-card {{
            background: {SURFACE};
            border: 1px solid {BORDER};
            border-radius: 12px;
            padding: 16px 18px 12px;
            height: 100%;
            min-height: 110px;
            box-sizing: border-box;
        }}
        .kpi-label {{
            font-size: 0.65rem;
            font-weight: 700;
            color: {MUTED};
            text-transform: uppercase;
            letter-spacing: 0.06em;
            margin-bottom: 5px;
        }}
        .kpi-value {{
            font-size: 1.7rem;
            font-weight: 800;
            color: {NAVY};
            line-height: 1.1;
        }}
        .kpi-value-sm {{
            font-size: 1.4rem;
            font-weight: 800;
            color: {NAVY};
            line-height: 1.1;
        }}
        .kpi-sub {{
            font-size: 0.7rem;
            color: {MUTED};
            margin-top: 3px;
        }}
        .kpi-pill {{
            display: inline-block;
            padding: 2px 9px;
            border-radius: 20px;
            font-size: 0.68rem;
            font-weight: 700;
            margin-top: 5px;
        }}
        .pill-green  {{ background:#D4EDDA; color:#1A6630; }}
        .pill-red    {{ background:#FAD7D4; color:#8B1A14; }}
        .pill-amber  {{ background:#FEF0CC; color:#7D4E00; }}
        .pill-blue   {{ background:{LIGHT_BLUE}; color:{NAVY}; }}

        /* ── Section headers ── */
        .sec-head {{
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 1rem;
            font-weight: 800;
            color: {NAVY};
            margin: 26px 0 12px 0;
            padding-bottom: 8px;
            border-bottom: 2px solid {BORDER};
        }}
        .sec-head span.badge {{
            background: {LIGHT_BLUE};
            color: {BLUE};
            border-radius: 6px;
            padding: 2px 9px;
            font-size: 0.65rem;
            font-weight: 700;
            letter-spacing: 0.05em;
        }}

        /* ── Filters bar ── */
        .filter-bar {{
            background: {SURFACE};
            border: 1px solid {BORDER};
            border-radius: 12px;
            padding: 16px 18px 10px;
            margin-bottom: 14px;
        }}
        .filter-label {{
            font-size: 0.62rem;
            font-weight: 700;
            color: {MUTED};
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 4px;
        }}

        /* ── Result badge (for table) ── */
        .badge-accept {{ background:#D4EDDA; color:#1A6630;
                         padding:2px 9px; border-radius:12px; font-size:0.78rem; font-weight:700; }}
        .badge-reject {{ background:#FAD7D4; color:#8B1A14;
                         padding:2px 9px; border-radius:12px; font-size:0.78rem; font-weight:700; }}
        .badge-other  {{ background:#E9ECEF; color:#495057;
                         padding:2px 9px; border-radius:12px; font-size:0.78rem; font-weight:700; }}

        /* ── Download button ── */
        div[data-testid="stDownloadButton"] > button {{
            background: {NAVY};
            color: white;
            border: none;
            border-radius: 8px;
            font-weight: 700;
            font-size: 0.9rem;
            padding: 10px 24px;
        }}
        div[data-testid="stDownloadButton"] > button:hover {{
            background: {BLUE};
        }}

        /* ── Dataframe tweaks ── */
        div[data-testid="stDataFrame"] {{ border-radius: 10px; overflow: hidden; }}

        /* ── Expander ── */
        div[data-testid="stExpander"] {{
            border: 1px solid {BORDER};
            border-radius: 10px;
        }}
    </style>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# Column detection & data cleaning
# ─────────────────────────────────────────────────────────────────────────────
COLUMN_KEYWORDS = {
    "demand_id": ["pmp", "demand id", "requisition", "req id", "position id"],
    "name":      ["candidate name", "name"],
    "pan":       ["pan number", "pan"],
    "vendor":    ["vendor"],
    "account":   ["account name", "account", "client"],
    "jrs":       ["jrs"],
    "status":    ["screen select reason", "screening status", "status", "screen result", "screen"],
    "remarks":   ["remarks or comments", "remarks", "comments", "reason"],
    "week":      ["screened date", "screened week", "week", "date"],
}


def normalize(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def find_col(columns, keywords):
    norm_map = {normalize(c): c for c in columns}
    for kw in sorted(keywords, key=len, reverse=True):
        nkw = normalize(kw)
        for ncol, orig in norm_map.items():
            if nkw == ncol:
                return orig
    for kw in sorted(keywords, key=len, reverse=True):
        nkw = normalize(kw)
        for ncol, orig in norm_map.items():
            if nkw in ncol:
                return orig
    return None


def classify_status(value):
    v = str(value).strip().lower()
    if any(k in v for k in ["select", "accept", "shortlist"]):
        return "Accepted"
    if any(k in v for k in ["reject", "decline"]):
        return "Rejected"
    return "Other"


def _find_header_row(xl, sheet, max_scan=6):
    """Detect the real header row (skips title/blank rows in formatted exports)."""
    all_kws = [kw for kws in COLUMN_KEYWORDS.values() for kw in kws]
    probe = xl.parse(sheet, header=None, nrows=max_scan)
    for idx, row in probe.iterrows():
        row_vals = [normalize(str(v)) for v in row if pd.notna(v)]
        for kw in all_kws:
            nkw = normalize(kw)
            if any(nkw == rv or nkw in rv for rv in row_vals):
                return int(idx)
    return 0


@st.cache_data(show_spinner=False)
def load_and_clean(file_bytes, file_name):
    xl = pd.ExcelFile(BytesIO(file_bytes))
    frames = []
    for sheet in xl.sheet_names:
        header_row = _find_header_row(xl, sheet)
        raw = xl.parse(sheet, header=header_row)
        if raw.empty or raw.shape[1] < 3:
            continue
        col_map = {field: find_col(raw.columns, kws) for field, kws in COLUMN_KEYWORDS.items()}
        if not col_map["vendor"] or not col_map["status"]:
            continue
        df = pd.DataFrame()
        for field, src_col in col_map.items():
            df[field] = raw[src_col] if  src_col else np.nan
        df = df.dropna(subset=["vendor"], how="any")
        df = df[df["vendor"].astype(str).str.strip() != ""]
        if df.empty:
            continue
        df["source_file"] = file_name
        frames.append(df)
    if not frames:
        return pd.DataFrame(columns=list(COLUMN_KEYWORDS) + ["source_file"])
    out = pd.concat(frames, ignore_index=True)
    out["jrs"]       = out["jrs"].fillna("Unspecified JRS").astype(str).str.strip().str.replace(r"\s*-\s*", "-", regex=True)
    out["vendor"]    = out["vendor"].astype(str).str.strip()
    out["demand_id"] = out["demand_id"].fillna("Unspecified").astype(str).str.strip()
    out["result"]    = out["status"].apply(classify_status)
    # Generate stable Candidate ID: CAND-{initials}-{4-char hash of normalised name}
    def _make_candidate_id(name):
        n = str(name).strip()
        parts = n.split()
        initials = "".join(p[0].upper() for p in parts if p)[:3]  # up to 3 initials
        h = hashlib.sha256(n.lower().encode()).hexdigest()[:4].upper()
        return f"CAND-{initials}-{h}"
    out["candidate_id"] = out["name"].apply(_make_candidate_id)
    # Convert date values in the 'week' column to ISO week labels (e.g. '2025-W03')
    def _date_to_week(val):
        if pd.isna(val):
            return np.nan
        try:
            dt = pd.to_datetime(val)
            return f"{dt.year}-W{dt.isocalendar().week:02d}"
        except Exception:
            return str(val).strip() or np.nan
    out["week"] = out["week"].apply(_date_to_week)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────
def build_excel_report(all_data, vendor_stats_display, top_vendors_display, best_vendor_per_jrs, weekly, has_week):
    WHITE_FONT  = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="0F2D5E")
    TITLE_FONT  = Font(bold=True, name="Arial", size=14, color="0F2D5E")
    BODY_FONT   = Font(name="Arial", size=10)
    THIN        = Side(style="thin", color="B7B7B7")
    BORDER_STYLE= Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER      = Alignment(horizontal="center", vertical="center")
    GREEN       = "C6EFCE"

    def write_df(ws, df, start_row=1, title=None, highlight_col=None):
        r = start_row
        if title:
            ws.cell(row=r, column=1, value=title).font = TITLE_FONT
            r += 2
        header_row = r
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=header_row, column=j, value=col)
            c.font, c.fill, c.alignment, c.border = WHITE_FONT, HEADER_FILL, CENTER, BORDER_STYLE
        for i, row in enumerate(df.itertuples(index=False), start=1):
            for j, val in enumerate(row, start=1):
                c = ws.cell(row=header_row + i, column=j, value=val)
                c.font, c.border = BODY_FONT, BORDER_STYLE
                c.alignment = CENTER if j > 1 else Alignment(horizontal="left")
                if highlight_col and df.columns[j - 1] == highlight_col:
                    c.fill = PatternFill("solid", fgColor=GREEN)
        for j, col in enumerate(df.columns, start=1):
            max_len = max((len(str(v)) for v in df[col]), default=0) if len(df) else 0
            width = max(len(str(col)), max_len) + 4
            ws.column_dimensions[get_column_letter(j)].width = min(max(width, 12), 55)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        return header_row + len(df) + 1

    total_demands    = all_data["demand_id"].nunique()
    total_candidates = len(all_data)
    total_accepted   = (all_data["result"] == "Accepted").sum()
    total_rejected   = (all_data["result"] == "Rejected").sum()
    total_other      = (all_data["result"] == "Other").sum()
    rate             = total_accepted / total_candidates if total_candidates else 0
    summary_df = pd.DataFrame(
        [
            ("Total Demands (unique PMP/Demand IDs)", total_demands),
            ("Total Unique JRS Roles",                all_data["jrs"].nunique()),
            ("Total Vendors Engaged",                 all_data["vendor"].nunique()),
            ("Total Candidates Screened",             total_candidates),
            ("Total Accepted",                        total_accepted),
            ("Total Rejected",                        total_rejected),
            ("Total Other/Pending",                   total_other),
            ("Overall Acceptance Rate",               f"{total_accepted}/{total_candidates} ({rate:.1%})"),
        ],
        columns=["Metric", "Value"],
    )

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    write_df(ws, summary_df, title="CV Screening — Overall Summary")

    ws2 = wb.create_sheet("Vendor Performance")
    ws2.sheet_view.showGridLines = False
    last_row = write_df(ws2, vendor_stats_display, title="Vendor Performance (All Vendors)")
    if len(vendor_stats_display):
        chart = BarChart()
        chart.title = "Acceptance Rate % by Vendor"
        chart.y_axis.title, chart.x_axis.title = "Acceptance Rate %", "Vendor"
        hr = 3
        data_ref = Reference(ws2, min_col=7, min_row=hr, max_row=hr + len(vendor_stats_display))
        cats_ref = Reference(ws2, min_col=1, min_row=hr + 1, max_row=hr + len(vendor_stats_display))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height, chart.width = 9, 18
        ws2.add_chart(chart, f"A{last_row + 2}")

    ws3 = wb.create_sheet("Top Vendors")
    ws3.sheet_view.showGridLines = False
    write_df(ws3, top_vendors_display, title="Top Vendors by Acceptance Rate", highlight_col="Rank")

    ws4 = wb.create_sheet("Best Vendor per JRS")
    ws4.sheet_view.showGridLines = False
    write_df(ws4, best_vendor_per_jrs, title="Best-Performing Vendor for Each JRS", highlight_col="Best Vendor")

    if has_week:
        from openpyxl.chart import LineChart
        ws5 = wb.create_sheet("Weekly Breakdown")
        ws5.sheet_view.showGridLines = False
        last_row_w = write_df(ws5, weekly, title="Weekly Screening Breakdown")
        if len(weekly):
            # header is on row 3 (title row 1, blank row 2, header row 3)
            hr_w = 3
            n_w  = len(weekly)
            wk_chart = LineChart()
            wk_chart.title  = "Weekly Screening Trend"
            wk_chart.y_axis.title = "Candidates"
            wk_chart.x_axis.title = "Week"
            wk_chart.style  = 10
            wk_chart.height = 12
            wk_chart.width  = 24
            # columns: Screened Week(1), Total_Screened(2), Accepted(3), Rejected(4), Acceptance Rate %(5)
            for col_idx, series_title in [(2, "Total Screened"), (3, "Accepted"), (4, "Rejected")]:
                data_ref = Reference(ws5, min_col=col_idx, min_row=hr_w, max_row=hr_w + n_w)
                wk_chart.add_data(data_ref, titles_from_data=True)
            cats_ref = Reference(ws5, min_col=1, min_row=hr_w + 1, max_row=hr_w + n_w)
            wk_chart.set_categories(cats_ref)
            ws5.add_chart(wk_chart, f"A{last_row_w + 2}")

    ws6 = wb.create_sheet("Raw Data")
    ws6.sheet_view.showGridLines = False
    raw_display = all_data.rename(columns={
        "demand_id": "Demand ID (PMP)", "candidate_id": "Candidate ID",
        "vendor": "Vendor", "account": "Account", "jrs": "JRS",
        "status": "Screening Status", "result": "Result",
        "remarks": "Remarks", "week": "Screened Week", "source_file": "Source File",
    })
    write_df(ws6, raw_display, title="Combined Raw Data")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# Helper: KPI card HTML
# ─────────────────────────────────────────────────────────────────────────────
def kpi_card(label, value, sub=None, pill=None, pill_cls="pill-blue"):
    pill_html = f'<div class="kpi-pill {pill_cls}">{pill}</div>' if pill else ""
    sub_html  = f'<div class="kpi-sub">{sub}</div>' if sub else ""
    return (
        f'<div class="kpi-card">'
        f'  <div class="kpi-label">{label}</div>'
        f'  <div class="kpi-value">{value}</div>'
        f'  {sub_html}{pill_html}'
        f'</div>'
    )


def sec_head(icon, title, badge=None):
    badge_html = f'<span class="badge">{badge}</span>' if badge else ""
    st.markdown(
        f'<div class="sec-head">{icon}&nbsp;{title} {badge_html}</div>',
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    # ── Logo / title ──
    st.markdown(
        """
        <div style="padding:10px 0 20px 0; border-bottom:1px solid #2A4A7F; margin-bottom:20px;">
            <div style="font-size:1.15rem; font-weight:800; color:#FFFFFF; letter-spacing:-0.01em;">
                📊 Hiring Dashboard
            </div>
            <div style="font-size:0.75rem; color:#8AAAD4; margin-top:2px;">CV Screening Analytics</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Data Source ──
    st.markdown(
        "<div style='font-size:0.65rem; font-weight:700; color:#8AAAD4; "
        "text-transform:uppercase; letter-spacing:0.07em; margin-bottom:8px;'>"
        "Data Source</div>",
        unsafe_allow_html=True,
    )
    uploaded_files = st.file_uploader(
        "Upload tracker file(s)",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="Upload one or more CV screening tracker Excel files.",
        label_visibility="collapsed",
    )
    if uploaded_files:
        names = ", ".join(f.name for f in uploaded_files)
        count_label = f"{len(uploaded_files)} file{'s' if len(uploaded_files) > 1 else ''} loaded"
        st.markdown(
            f"""
            <div style="background:#1A3D6E; border:1px dashed #3A6AAE; border-radius:8px;
                        padding:10px 12px; color:#B8CFEE; font-size:0.72rem; text-align:center;
                        margin-top:4px; margin-bottom:6px; word-break:break-all;">
                📂 {names}<br>
                <span style="color:#5A7AAA; font-size:0.65rem;">{count_label}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Ranking Settings ──
    st.markdown(
        "<div style='font-size:0.65rem; font-weight:700; color:#8AAAD4; "
        "text-transform:uppercase; letter-spacing:0.07em; margin:18px 0 8px;'>"
        "Ranking Settings</div>",
        unsafe_allow_html=True,
    )
    min_submissions = st.slider("Min. submissions to rank vendor", 1, 10, 2)
    top_n           = st.slider("Top N vendors to highlight",      3, 10, 5)

    st.markdown(
        "<div style='margin-top:32px; padding-top:16px; border-top:1px solid #2A4A7F; "
        "font-size:0.65rem; color:#5A7AAA; text-align:center;'>IBM Hiring Analytics · v2.0</div>",
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────────────────────────────────────
# Landing state — no file uploaded
# ─────────────────────────────────────────────────────────────────────────────
if not uploaded_files:
    st.markdown(
        f"""
        <div class="dash-banner">
            <h1>Hiring Assistance Dashboard</h1>
            <p>CV Screening · Vendor Performance · Demand Analytics</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.info(
        "👈 **Get started** — upload one or more CV screening tracker Excel files "
        "using the sidebar on the left.",
        icon="📂",
    )
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# Load data
# ─────────────────────────────────────────────────────────────────────────────
with st.spinner("Loading and processing data…"):
    all_data = pd.concat(
        [load_and_clean(f.getvalue(), f.name) for f in uploaded_files],
        ignore_index=True,
    )

if all_data.empty:
    st.error(
        "**No usable candidate rows found.** "
        "Check that your file has Vendor and Status columns.",
        icon="⚠️",
    )
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# Banner
# ─────────────────────────────────────────────────────────────────────────────
file_label = uploaded_files[0].name if len(uploaded_files) == 1 else f"{len(uploaded_files)} files loaded"
st.markdown(
    f"""
    <div class="dash-banner">
        <h1>Hiring Assistance Dashboard</h1>
        <p>CV Screening · Vendor Performance · Demand Analytics &nbsp;·&nbsp;
           <span style="color:#7FAADD;">{file_label}</span></p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# KPI strip — with week filter
# ─────────────────────────────────────────────────────────────────────────────

# Build week-ending options (reused helper, same logic as drill-down filter)
def _week_str_to_friday(w):
    try:
        monday = pd.to_datetime(str(w) + "-1", format="%G-W%V-%u")
        return monday + pd.Timedelta(days=4)
    except Exception:
        return pd.NaT

_raw_weeks   = all_data["week"].dropna().unique()
_friday_map  = {}
for _w in _raw_weeks:
    _fri = _week_str_to_friday(str(_w))
    if pd.notna(_fri):
        _friday_map[_fri.strftime("%d %b %Y").lstrip("0")] = str(_w)

# ── Week 0 — static summary numbers (43 demands, 102 candidates, 71 selected, 31 rejected)
# Build a minimal dataframe: 71 "Screen Select" rows + 31 "Screen Reject" rows across 43 demand IDs
_w0_demands  = [f"W0-DEMAND-{i+1:02d}" for i in range(43)]
_w0_statuses = (["Screen Select"] * 71) + (["Screen Reject"] * 31)
_w0_dem_ids  = [_w0_demands[i % 43] for i in range(102)]
WEEK0_DATA = pd.DataFrame({
    "demand_id":    _w0_dem_ids,
    "candidate_id": [f"CAND-W0-{i+1:03d}" for i in range(102)],
    "name":         [f"Candidate {i+1}" for i in range(102)],
    "vendor":       "—",
    "jrs":          "—",
    "status":       _w0_statuses,
    "result":       [classify_status(s) for s in _w0_statuses],
    "remarks":      "—",
    "week":         "Week 0",
    "account":      "—",
    "source_file":  "Week 0",
})
# ─────────────────────────────────────────────────────────────────────────────

_kpi_week_options = ["All"] + ["Week 0"] + sorted(_friday_map.keys(), key=lambda l: _friday_map[l])

# Read persisted week selection (default "Week 0" on first load)
_kpi_week_default = st.session_state.get("kpi_week_sel", "All")
if _kpi_week_default not in _kpi_week_options:
    _kpi_week_default = "All"

if _kpi_week_default == "Week 0":
    kpi_df = WEEK0_DATA
elif _kpi_week_default == "All":
    kpi_df = all_data
else:
    kpi_df = all_data[all_data["week"] == _friday_map.get(_kpi_week_default, "")]

total_demands    = kpi_df["demand_id"].nunique()
total_candidates = len(kpi_df)
total_accepted   = int((kpi_df["result"] == "Accepted").sum())
total_rejected   = int((kpi_df["result"] == "Rejected").sum())
total_other      = int((kpi_df["result"] == "Other").sum())
overall_rate     = total_accepted / total_candidates if total_candidates else 0
n_vendors        = kpi_df["vendor"].nunique()
n_jrs            = kpi_df["jrs"].nunique()

_is_week0 = (_kpi_week_default == "Week 0")

# Row 1 — 4 main KPI cards (equal width, full row)
k1, k2, k3, k4 = st.columns(4)
with k1:
    # hide JRS pill when Week 0 (no JRS data)
    st.markdown(kpi_card("Demands", total_demands,
                         pill=None if _is_week0 else f"{n_jrs} JRS roles",
                         pill_cls="pill-blue"), unsafe_allow_html=True)
with k2:
    st.markdown(kpi_card("Candidates Screened", f"{total_candidates:,}",
                         sub=f"Across {total_demands} demand IDs"), unsafe_allow_html=True)
with k3:
    st.markdown(kpi_card("Selected", f"{total_accepted:,}",
                         pill=f"{overall_rate:.1%} acceptance rate", pill_cls="pill-green"), unsafe_allow_html=True)
with k4:
    st.markdown(kpi_card("Rejected", f"{total_rejected:,}",
                         pill=f"{total_rejected/total_candidates:.1%} of total" if total_candidates else "—",
                         pill_cls="pill-red"), unsafe_allow_html=True)

# Row 2 — hide Vendors Engaged when Week 0 (no vendor data); always show Week Ending filter
st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)
if _is_week0:
    # Only the filter, full width
    _, k6 = st.columns([3, 1])
else:
    k5, k6 = st.columns(2)
    with k5:
        st.markdown(
            f'<div class="kpi-card" style="padding:10px 16px;">'
            f'  <div class="kpi-label">Vendors Engaged</div>'
            f'  <div class="kpi-value-sm" style="margin-top:6px;">{n_vendors}</div>'
            f'  <div class="kpi-sub" style="margin-top:4px;">Min. {min_submissions} CVs threshold</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
with k6:
    st.markdown('<div class="filter-label">Week Ending</div>', unsafe_allow_html=True)
    st.selectbox(
        "KPI Week Ending", _kpi_week_options,
        index=_kpi_week_options.index(_kpi_week_default),
        key="kpi_week_sel",
        label_visibility="collapsed",
    )

st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Section 1 — Demand Drill-down
# ─────────────────────────────────────────────────────────────────────────────
sec_head("🔍", "Demand Drill-down")

with st.container():
    if _is_week0:
        # Week 0 has no JRS/vendor — show only Demand + Screening Status filters
        _src_df = WEEK0_DATA
        f1, f2 = st.columns(2)
        with f1:
            st.markdown('<div class="filter-label">Demand ID</div>', unsafe_allow_html=True)
            demand_options  = ["All"] + sorted(_src_df["demand_id"].dropna().unique().tolist())
            selected_demand = st.selectbox("Demand ID", demand_options, label_visibility="collapsed")
        demand_df  = _src_df if selected_demand == "All" else _src_df[_src_df["demand_id"] == selected_demand]
        jrs_df     = demand_df
        vendor_df  = demand_df
        with f2:
            st.markdown('<div class="filter-label">Screening Status</div>', unsafe_allow_html=True)
            selected_result = st.selectbox(
                "Screening Status",
                ["All", "Screen Select", "Screen Reject"],
                label_visibility="collapsed",
            )
        status_df  = demand_df if selected_result == "All" else demand_df[demand_df["status"].str.strip() == selected_result]
        filtered_df = status_df
        selected_week = "All"   # no week sub-filter for Week 0
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        f1, f2, f3, f4, f5 = st.columns(5)
        with f1:
            st.markdown('<div class="filter-label">Demand ID</div>', unsafe_allow_html=True)
            demand_options  = ["All"] + sorted(all_data["demand_id"].unique().tolist())
            selected_demand = st.selectbox("Demand ID", demand_options, label_visibility="collapsed")
        demand_df = all_data if selected_demand == "All" else all_data[all_data["demand_id"] == selected_demand]

        with f2:
            st.markdown('<div class="filter-label">JRS Role</div>', unsafe_allow_html=True)
            jrs_options  = ["All"] + sorted(demand_df["jrs"].unique().tolist())
            selected_jrs = st.selectbox("JRS Role", jrs_options, label_visibility="collapsed")
        jrs_df = demand_df if selected_jrs == "All" else demand_df[demand_df["jrs"] == selected_jrs]

        with f3:
            st.markdown('<div class="filter-label">Vendor</div>', unsafe_allow_html=True)
            vendor_options  = ["All"] + sorted(jrs_df["vendor"].unique().tolist())
            selected_vendor = st.selectbox("Vendor", vendor_options, label_visibility="collapsed")
        vendor_df = jrs_df if selected_vendor == "All" else jrs_df[jrs_df["vendor"] == selected_vendor]

        with f4:
            st.markdown('<div class="filter-label">Screening Status</div>', unsafe_allow_html=True)
            selected_result = st.selectbox(
                "Screening Status",
                ["All", "Screen Select", "Screen Reject"],
                label_visibility="collapsed",
            )
        status_df = vendor_df if selected_result == "All" else vendor_df[vendor_df["status"].str.strip() == selected_result]

        with f5:
            st.markdown('<div class="filter-label">Week Ending (Friday)</div>', unsafe_allow_html=True)
            raw_weeks  = all_data["week"].dropna().unique()
            friday_map = {}
            for w in raw_weeks:
                _fri2 = _week_str_to_friday(str(w))
                if pd.notna(_fri2):
                    friday_map[_fri2.strftime("%d %b %Y").lstrip("0")] = str(w)
            week_options  = ["All"] + sorted(friday_map.keys(), key=lambda lbl: friday_map[lbl])
            selected_week = st.selectbox("Week Ending", week_options, label_visibility="collapsed")

        if selected_week == "All":
            filtered_df = status_df
        else:
            filtered_df = status_df[status_df["week"] == friday_map[selected_week]]
        st.markdown('</div>', unsafe_allow_html=True)

# Drill-down KPI row — uses smaller value font (kpi-value-sm)
n_acc = int((filtered_df["result"] == "Accepted").sum())
n_rej = int((filtered_df["result"] == "Rejected").sum())
n_oth = int((filtered_df["result"] == "Other").sum())
acc_r = n_acc / len(filtered_df) if len(filtered_df) else 0

def kpi_card_sm(label, value, sub=None, pill=None, pill_cls="pill-blue"):
    pill_html = f'<div class="kpi-pill {pill_cls}">{pill}</div>' if pill else ""
    sub_html  = f'<div class="kpi-sub">{sub}</div>' if sub else ""
    return (
        f'<div class="kpi-card">'
        f'  <div class="kpi-label">{label}</div>'
        f'  <div class="kpi-value-sm">{value}</div>'
        f'  {sub_html}{pill_html}'
        f'</div>'
    )

if _is_week0:
    # Week 0: clean 3-card row — Candidates, Selected, Rejected only
    dm1, dm2, dm3 = st.columns(3)
    with dm1:
        st.markdown(kpi_card_sm("Candidates in Selection", f"{len(filtered_df):,}"), unsafe_allow_html=True)
    with dm2:
        st.markdown(kpi_card_sm("Selected (filtered)", n_acc,
                                pill=f"{acc_r:.0%}", pill_cls="pill-green"), unsafe_allow_html=True)
    with dm3:
        st.markdown(kpi_card_sm("Rejected (filtered)", n_rej,
                                pill_cls="pill-red",
                                pill=f"{n_rej/len(filtered_df):.0%}" if len(filtered_df) else "—"), unsafe_allow_html=True)
else:
    dm1, dm2, dm3, dm4, dm5 = st.columns(5)
    with dm1:
        st.markdown(kpi_card_sm("Candidates in Selection", f"{len(filtered_df):,}"), unsafe_allow_html=True)
    with dm2:
        st.markdown(kpi_card_sm("JRS Roles in Selection", filtered_df["jrs"].nunique()), unsafe_allow_html=True)
    with dm3:
        st.markdown(kpi_card_sm("Accepted (filtered)", n_acc,
                                pill=f"{acc_r:.0%}", pill_cls="pill-green"), unsafe_allow_html=True)
    with dm4:
        st.markdown(kpi_card_sm("Rejected (filtered)", n_rej,
                                pill_cls="pill-red",
                                pill=f"{n_rej/len(filtered_df):.0%}" if len(filtered_df) else "—"), unsafe_allow_html=True)
    with dm5:
        st.markdown(kpi_card_sm("Pending / Other", n_oth,
                                pill_cls="pill-amber",
                                pill=f"{n_oth/len(filtered_df):.0%}" if len(filtered_df) else "—"), unsafe_allow_html=True)

st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)

# Candidate table (name replaced with stable ID; Result column excluded from UI but kept in Excel)
display_cols = {
    "demand_id": "Demand ID", "candidate_id": "Candidate ID", "jrs": "JRS",
    "vendor": "Vendor", "status": "Screening Status", "remarks": "Remarks",
}
available_cols = [c for c in display_cols if c in filtered_df.columns]
table_df = filtered_df[available_cols].rename(columns=display_cols).reset_index(drop=True)

st.dataframe(
    table_df,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Result": st.column_config.TextColumn("Result"),
        "Demand ID": st.column_config.TextColumn("Demand ID"),
        "Candidate Name": st.column_config.TextColumn("Candidate Name", width="medium"),
    },
)
st.caption(f"{len(filtered_df):,} candidate record(s) shown · filtered from {total_candidates:,} total")

# ─────────────────────────────────────────────────────────────────────────────
# Section 2 — Vendor Performance
# ─────────────────────────────────────────────────────────────────────────────
sec_head("🏢", "Vendor Performance", badge=f"All {n_vendors} vendors")

vendor_stats = (
    all_data.groupby("vendor")
    .agg(
        Total_Submitted=("result", "count"),
        Accepted=("result",  lambda s: (s == "Accepted").sum()),
        Rejected=("result",  lambda s: (s == "Rejected").sum()),
        Other=("result",     lambda s: (s == "Other").sum()),
    )
    .reset_index()
    .rename(columns={"vendor": "Vendor"})
)
vendor_stats["Acceptance Rate %"] = (
    vendor_stats["Accepted"] / vendor_stats["Total_Submitted"] * 100
).round(1)
vendor_stats["Acceptance Rate"] = vendor_stats.apply(
    lambda r: f"{r['Accepted']}/{r['Total_Submitted']}", axis=1
)
vendor_stats = vendor_stats.sort_values("Acceptance Rate %", ascending=False).reset_index(drop=True)
vendor_stats["Meets Threshold"] = vendor_stats["Total_Submitted"] >= min_submissions

top_vendors = vendor_stats.sort_values(
    ["Meets Threshold", "Acceptance Rate %", "Total_Submitted"], ascending=[False, False, False]
).head(top_n).reset_index(drop=True)
top_vendors.insert(0, "Rank", range(1, len(top_vendors) + 1))

vendor_stats_display = vendor_stats[
    ["Vendor", "Total_Submitted", "Accepted", "Rejected", "Other", "Acceptance Rate", "Acceptance Rate %"]
]
top_vendors_display = top_vendors[
    ["Rank", "Vendor", "Total_Submitted", "Accepted", "Acceptance Rate", "Acceptance Rate %"]
]

# Bar chart — full width but compact height
fig_bar = px.bar(
    vendor_stats_display.sort_values("Acceptance Rate %"),
    x="Acceptance Rate %", y="Vendor",
    orientation="h",
    text="Acceptance Rate",
    color="Acceptance Rate %",
    color_continuous_scale=["#F4CCCC", "#FFE699", "#C6EFCE"],
    labels={"Acceptance Rate %": "Acceptance Rate (%)"},
)
fig_bar.update_traces(textposition="outside", textfont_size=10)
fig_bar.update_layout(
    height=min(max(220, len(vendor_stats_display) * 28), 360),
    margin=dict(t=10, b=10, l=10, r=50),
    coloraxis_showscale=False,
    plot_bgcolor="white",
    paper_bgcolor="white",
    yaxis=dict(tickfont=dict(size=10)),
    xaxis=dict(gridcolor="#EEEEEE", zeroline=False, tickfont=dict(size=10)),
)
st.plotly_chart(fig_bar, use_container_width=True)

# Row 2 — donut (left) + top vendors table (right)
vc2, vc3 = st.columns([1, 2])

with vc2:
    donut_df = pd.DataFrame({
        "Result": ["Accepted", "Rejected", "Other"],
        "Count":  [total_accepted, total_rejected, total_other],
    })
    fig_donut = go.Figure(go.Pie(
        labels=donut_df["Result"],
        values=donut_df["Count"],
        hole=0.58,
        marker_colors=[TEAL, RED, AMBER],
        textinfo="percent",
        textfont_size=12,
        hovertemplate="%{label}: %{value} (%{percent})<extra></extra>",
    ))
    fig_donut.update_layout(
        title=dict(text="Overall Outcome Mix", font=dict(size=13, color=NAVY), x=0.5),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5, font=dict(size=11)),
        margin=dict(t=40, b=20, l=10, r=10),
        height=320,
        paper_bgcolor="white",
    )
    fig_donut.add_annotation(
        text=f"<b>{overall_rate:.0%}</b><br><span style='font-size:10px'>Accept</span>",
        x=0.5, y=0.5, showarrow=False,
        font=dict(size=16, color=NAVY),
        align="center",
    )
    st.plotly_chart(fig_donut, use_container_width=True)

with vc3:
    st.markdown(f"<div style='font-size:0.8rem; font-weight:700; color:{NAVY}; margin-bottom:8px;'>🏆 Top {top_n} Vendors</div>", unsafe_allow_html=True)
    st.dataframe(top_vendors_display, use_container_width=True, hide_index=True,
                 height=35 * (len(top_vendors_display) + 1) + 3)

with st.expander("View full vendor performance table"):
    st.dataframe(vendor_stats_display, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# Section 3 — Best Vendor per JRS
# ─────────────────────────────────────────────────────────────────────────────
sec_head("🎯", "Best Vendor per JRS Role", badge=f"{all_data['jrs'].nunique()} roles")

jrs_vendor = (
    all_data.groupby(["jrs", "vendor"])
    .agg(
        Total_Submitted=("result", "count"),
        Accepted=("result", lambda s: (s == "Accepted").sum()),
    )
    .reset_index()
)
jrs_vendor["Acceptance Rate %"] = (
    jrs_vendor["Accepted"] / jrs_vendor["Total_Submitted"] * 100
).round(1)
eligible_jv = jrs_vendor[jrs_vendor["Total_Submitted"] >= min_submissions]

best_rows = []
for jrs, grp in jrs_vendor.groupby("jrs"):
    pool = eligible_jv[eligible_jv["jrs"] == jrs]
    if pool.empty:
        pool = grp
    best = pool.sort_values(
        ["Acceptance Rate %", "Total_Submitted", "Accepted"], ascending=[False, False, False]
    ).iloc[0]
    best_rows.append(best)

best_vendor_per_jrs = (
    pd.DataFrame(best_rows)
    .rename(columns={"jrs": "JRS", "vendor": "Best Vendor"})
    .reset_index(drop=True)
)
best_vendor_per_jrs["Acceptance Rate"] = best_vendor_per_jrs.apply(
    lambda r: f"{int(r['Accepted'])}/{int(r['Total_Submitted'])}", axis=1
)
best_vendor_per_jrs = best_vendor_per_jrs[
    ["JRS", "Best Vendor", "Total_Submitted", "Accepted", "Acceptance Rate", "Acceptance Rate %"]
]

bv1, bv2 = st.columns([3, 2])
with bv1:
    st.dataframe(best_vendor_per_jrs, use_container_width=True, hide_index=True)
with bv2:
    fig_jrs = px.bar(
        best_vendor_per_jrs.sort_values("Acceptance Rate %"),
        x="Acceptance Rate %", y="JRS",
        orientation="h",
        color="Best Vendor",
        text="Best Vendor",
        color_discrete_sequence=PALETTE,
    )
    fig_jrs.update_traces(textposition="inside", textfont_size=10, insidetextanchor="middle")
    fig_jrs.update_layout(
        height=max(260, len(best_vendor_per_jrs) * 42),
        margin=dict(t=10, b=10, l=10, r=10),
        showlegend=False,
        plot_bgcolor="white",
        paper_bgcolor="white",
        xaxis=dict(gridcolor="#EEEEEE"),
    )
    st.plotly_chart(fig_jrs, use_container_width=True)

# ─────────────────────────────────────────────────────────────────────────────
# Section 4 — Weekly Trend (conditional)
# ─────────────────────────────────────────────────────────────────────────────
has_week = all_data["week"].notna().any()
weekly   = pd.DataFrame(columns=["Screened Week", "Total_Screened", "Accepted", "Rejected", "Acceptance Rate %"])

if has_week:
    sec_head("📅", "Weekly Screening Trend")
    weekly = (
        all_data.dropna(subset=["week"])
        .groupby("week")
        .agg(
            Total_Screened=("result", "count"),
            Accepted=("result",  lambda s: (s == "Accepted").sum()),
            Rejected=("result",  lambda s: (s == "Rejected").sum()),
        )
        .reset_index()
    )
    weekly["Acceptance Rate %"] = (weekly["Accepted"] / weekly["Total_Screened"] * 100).round(1)
    # Convert ISO week string (e.g. "2025-W03") → Friday of that week (e.g. "17 Jan 2025")
    def _week_to_ending(w):
        try:
            monday = pd.to_datetime(str(w) + "-1", format="%G-W%V-%u")
            return (monday + pd.Timedelta(days=4)).strftime("%d %b %Y").lstrip("0")
        except Exception:
            return str(w)
    weekly["Screened Week"] = weekly["week"].apply(_week_to_ending)
    # Sort by original ISO week so chart order is chronological
    weekly = (weekly.sort_values("week")
              .drop(columns=["week"])
              .reset_index(drop=True)
              [["Screened Week", "Total_Screened", "Accepted", "Rejected", "Acceptance Rate %"]])

    fig_week = go.Figure()
    fig_week.add_trace(go.Scatter(
        x=weekly["Screened Week"], y=weekly["Total_Screened"],
        name="Total Screened", mode="lines+markers",
        line=dict(color=BLUE, width=2.5), marker=dict(size=7),
    ))
    fig_week.add_trace(go.Scatter(
        x=weekly["Screened Week"], y=weekly["Accepted"],
        name="Accepted", mode="lines+markers",
        line=dict(color=TEAL, width=2.5, dash="dot"), marker=dict(size=7),
    ))
    fig_week.add_trace(go.Scatter(
        x=weekly["Screened Week"], y=weekly["Rejected"],
        name="Rejected", mode="lines+markers",
        line=dict(color=RED, width=2.5, dash="dot"), marker=dict(size=7),
    ))
    fig_week.update_layout(
        height=340,
        margin=dict(t=10, b=10),
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        xaxis=dict(gridcolor="#EEEEEE"),
        yaxis=dict(gridcolor="#EEEEEE"),
    )
    st.plotly_chart(fig_week, use_container_width=True)

    with st.expander("View weekly data table"):
        st.dataframe(weekly, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# Section 5 — Export
# ─────────────────────────────────────────────────────────────────────────────
sec_head("⬇️", "Export Report")

exp_col, _ = st.columns([2, 5])
with exp_col:
    excel_buf = build_excel_report(
        all_data, vendor_stats_display, top_vendors_display,
        best_vendor_per_jrs, weekly, has_week,
    )
    st.download_button(
        "⬇️  Download Full Excel Report",
        data=excel_buf,
        file_name="CV_Screening_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("Includes Summary, Vendor Performance, Top Vendors, Best Vendor per JRS, and Raw Data sheets.")
